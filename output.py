import os

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side

dict_rows_header = {'x1': 'Atuador estático', 'x2': 'Atuador de rotação', 'x3': 'Atuador pneumático',
                    'x4': 'Base T galvanizada', 'x5': 'Válvula borboleta', 'x6': 'Interruptor deslizante',
                    'x7': 'Interruptor de tecla', 'x8': 'Inversor', 'x9': 'Chicote 7 vias', 'x10': 'Chicote 4 vias',
                    'x11': 'Comutador duplo', 'x12': 'Comutador inversor', 'x13': 'Comutador de ignição',
                    'x14': 'Válvula Bipartida', 'x15': 'Válvula Monobloco', 'x16': 'Junta de expansão',
                    'x17': 'Base X galvanizada', 'x18': 'Microrutor', 'x19': 'Chave micro switch',
                    'x20': 'Terminal 2 eixos', 'x21': 'Terminal 4 eixos', 'x22': 'Terminal 6 eixos',
                    'x23': 'Sensor de pressão', 'x24': 'Sensor de aproximação', 'x25': 'Flexinity',
                    'x26': 'Flexinity Mini', 'x27': 'Oscilador', 'x28': 'Timer', 'x29': 'Termistor',
                    'x30': 'Potenciômetro deslizante', 'Z': 'Objective'}


def export_to_sheet(columns_header: list, rows_header: list, matrix: np.array, last_row_filled: int,
                    columns_to_color=None, row_to_color=None) -> None:
    file_name = "simplex.xlsx"
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active

    color_fill = PatternFill(start_color='5F9F9F', end_color='5F9F9F', fill_type='solid')
    border_style = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                          top=Side(border_style='thin'), bottom=Side(border_style='thin'))

    for i, column in enumerate(columns_header):
        worksheet.cell(row=last_row_filled, column=i + 2, value=column)

    for j, row in enumerate(rows_header):
        worksheet.cell(row=last_row_filled + j + 1, column=1, value=row)

    for a, row in enumerate(matrix):
        for b, column in enumerate(row):
            worksheet.cell(row=last_row_filled + a + 1, column=b + 2, value=column)

    if columns_to_color is not None:
        for col in range(1, len(columns_header) + 2):
            worksheet.cell(row=last_row_filled + row_to_color + 1, column=col).fill = color_fill

        for row in range(last_row_filled, last_row_filled + len(rows_header) + 1):
            worksheet.cell(row=row, column=columns_to_color + 2).fill = color_fill

    for row in worksheet.iter_rows(min_row=last_row_filled, max_row=last_row_filled + len(rows_header), min_col=1,
                                   max_col=len(columns_header) + 1):
        for cell in row:
            cell.border = border_style

    workbook.save(file_name)


def export_results_to_txt(rows_headers: list, matrix: np.array, number_of_iterations: int):
    with open('results.txt', 'w') as file:
        for i, row in enumerate(rows_headers):
            if row in dict_rows_header.keys():
                file.write(f'{dict_rows_header[row]} : {matrix[i][-1]} \n')
        file.write(f'Number of iterations : {number_of_iterations}')
